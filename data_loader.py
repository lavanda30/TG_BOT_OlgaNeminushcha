"""
data_loader.py — спільний модуль завантаження прайсів з all_products.xlsx

Підтримує:
  - локальний файл (EXCEL_PATH)
  - URL (EXCEL_URL) — для Railway або будь-якого хостингу

Повертає dict: { supplier_name: [ {row_dict}, ... ] }
"""
import os
import re
import io
import logging
import urllib.request

logger = logging.getLogger(__name__)

COLS = [
    'supplier', 'sku', 'name', 'category', 'fabric',
    'color', 'width_cm', 'height_cm',
    'price', 'price_retail',
    'currency', 'unit', 'in_stock', 'collection', 'contacts'
]

UAH_RATE = 45


def _load_workbook_bytes(data: bytes):
    try:
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        logger.error(f"openpyxl error: {e}")
        raise


def _fetch_excel() -> bytes:
    url = os.environ.get("EXCEL_URL", "").strip()
    path = os.environ.get("EXCEL_PATH", "").strip()

    if url:
        # Автоматично конвертуємо GitHub blob → raw
        if "github.com" in url and "/blob/" in url:
            url = url.replace("github.com", "raw.githubusercontent.com").replace("/blob/", "/")
            logger.info(f"Converted to raw GitHub URL: {url}")
        logger.info(f"Downloading Excel from: {url}")
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            return r.read()

    if path:
        logger.info(f"Reading Excel from: {path}")
        with open(path, "rb") as f:
            return f.read()

    for candidate in [
        os.path.join(os.path.dirname(__file__), "..", "all_products.xlsx"),
        os.path.join(os.path.dirname(__file__), "all_products.xlsx"),
        "all_products.xlsx",
    ]:
        if os.path.exists(candidate):
            logger.info(f"Found Excel at: {candidate}")
            with open(candidate, "rb") as f:
                return f.read()

    raise FileNotFoundError(
        "all_products.xlsx не знайдено. "
        "Встановіть EXCEL_URL або EXCEL_PATH."
    )


def load_all(allowed_suppliers: list | None = None) -> dict:
    data = _fetch_excel()
    wb = _load_workbook_bytes(data)

    result = {}

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            d = {}
            for i, col in enumerate(COLS):
                d[col] = row[i] if i < len(row) else None

            supplier = str(d['supplier']).strip() if d['supplier'] else None
            if not supplier:
                continue

            sku = str(d['sku']).strip() if d['sku'] else ''
            if 'КОНТАКТИ' in sku.upper():
                continue

            if allowed_suppliers is not None:
                if supplier not in allowed_suppliers:
                    continue

            if supplier not in result:
                result[supplier] = []
            result[supplier].append(d)

    logger.info(
        f"Loaded {sum(len(v) for v in result.values())} rows "
        f"across {len(result)} suppliers"
    )
    return result


def fmt_price(row: dict) -> str:
    currency = str(row.get('currency') or 'USD').strip()
    retail = row.get('price_retail')
    price = row.get('price')

    if retail is None and price is None:
        return "—"

    main = retail if retail is not None else price

    try:
        main_f = float(main)
    except (TypeError, ValueError):
        return str(main)

    if currency.upper() in ('USD', 'У.Е.', 'U.E.', '$'):
        uah = round(main_f * UAH_RATE * 2)
        val = int(main_f) if main_f == int(main_f) else main_f
        return f"*{val}$* · ~{uah}грн"
    else:
        if price is not None and retail is not None and price != retail:
            try:
                p = float(price)
                r = float(retail)
                return f"*{p:.2f}* / *{r:.2f}* грн"
            except Exception:
                pass
        return f"*{main_f:.2f}* грн"


def get_tag(row: dict) -> str:
    in_stock = str(row.get('in_stock') or '').upper()
    coll = str(row.get('collection') or '').upper()

    if 'OUT OF STOCK' in in_stock or 'ЗНЯТО' in in_stock:
        return '⛔'
    if 'SALE' in in_stock or 'РОЗПРОДАЖ' in in_stock:
        return '🔴'
    if 'ORDER' in in_stock or 'ЗАМОВЛЕННЯ' in coll:
        return '📦'
    if 'НОВИНКА' in coll:
        return '🟢'
    if 'ЗНИЖЕННЯ' in coll or 'ЗНИЖЕНА' in coll:
        return '🟡'
    return ''


def get_extra(row: dict) -> str:
    for field in ('fabric', 'category'):
        val = row.get(field)
        if val and str(val).strip() not in ('', '—', 'None'):
            v = str(val).strip()
            if len(v) > 60:
                v = v[:57] + '...'
            return v
    return ''


def normalize(s: str) -> str:
    return re.sub(r'[\s\-_/.]', '', str(s)).lower()
